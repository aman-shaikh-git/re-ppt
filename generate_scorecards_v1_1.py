import win32com.client
import os
import time
import re
from pathlib import Path
from openpyxl import load_workbook
import pythoncom



def excel_to_pptx_text(val):
    """
    Standardizes text from Excel for PowerPoint.
    Converts Excel newlines (\n) to PPT carriage returns (\r).
    """
    if val is None: return ""
    text = str(val)
    # Convert Excel line breaks to PPT line breaks
    text = text.replace('\n', '\r')
    # Clean out any other illegal ASCII control characters
    illegal_chars = re.compile(r'[\000-\010\013\014\016-\037]')
    return illegal_chars.sub("", text)

def update_table_v1(table_obj, raw_data):
    """
    Improved Table Logic:
    Rows separated by '||'
    Columns separated by '|'
    Cells can contain standard newlines.
    """
    if not raw_data: return
    
    # Split by double pipe for rows
    row_strings = [r.strip() for r in str(raw_data).split('||') if r.strip()]
    
    for r_idx, row_str in enumerate(row_strings):
        if r_idx >= table_obj.Rows.Count: break
        
        # Split by single pipe for columns
        col_strings = [c.strip() for c in row_str.split('|')]
        
        for c_idx, cell_text in enumerate(col_strings):
            if c_idx >= table_obj.Columns.Count: break
            
            # Format the cell text (handle newlines inside the cell)
            formatted_text = excel_to_pptx_text(cell_text)
            table_obj.Cell(r_idx + 1, c_idx + 1).Shape.TextFrame.TextRange.Text = formatted_text

def generate_deck(pptx_path, excel_path, output_path):
    """
    Refactored core logic.
    Takes two paths, returns the path to the final PPTX.
    """
    pythoncom.CoInitialize()
    pptx_path = Path(pptx_path)
    excel_path = Path(excel_path)
    #output_path = pptx_path.parent / (pptx_path.stem + "_generated.pptx") #Deprecated to allow output to be user input

    # Load Data
    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb["Content"]
    tags = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    max_col = ws.max_column


# Start PowerPoint
    try:
        ppt_app = win32com.client.GetActiveObject("PowerPoint.Application")
    except:
        ppt_app = win32com.client.Dispatch("PowerPoint.Application")
    
    prs = ppt_app.Presentations.Open(str(pptx_path.absolute()), WithWindow=True)
    source_slide = prs.Slides(1)

 # Track the insertion point (start after the template)
    current_pos = 2 
    slide_counter = 0

    # Generation Loop
    for col_idx in range(3, max_col + 1):
        scorecard_name = ws.cell(row=1, column=col_idx).value
        print(f"Generating: {scorecard_name}")

        new_slide = source_slide.Duplicate()
        # Ensure slide is moved to the end to maintain Excel order
        new_slide.MoveTo(prs.Slides.Count)
        
        # Collect column data
        column_data = {}
        for row_idx, tag in enumerate(tags, start=2):
            column_data[tag] = ws.cell(row=row_idx, column=col_idx).value

        # Update Shapes
        for shape in new_slide.Shapes:
            # Tag Matching Logic (Simplified)
            s_name = shape.Name.replace(" ", "_")
            if shape.HasTable: kind = "Table"
            elif shape.HasTextFrame:
                kind = "Label" if len(shape.TextFrame.TextRange.Text.strip()) < 25 else "TextBox"
            else: kind = "Shape"
            tag = f"{kind}_{s_name}"

            if tag in column_data and column_data[tag] is not None:
                data_val = column_data[tag]
                
                try:
                    if shape.HasTextFrame:
                        shape.TextFrame.TextRange.Text = excel_to_pptx_text(data_val)
                    elif shape.HasTable:
                        update_table_v1(shape.Table, data_val)
                except Exception as e:
                    print(f"  Error on {tag}: {e}")
            slide_counter+=1
    
    abs_path = str(Path(output_path).absolute())
    prs.SaveAs(str(abs_path))
    # We don't close the presentation so the user can see it, 
    # but we return the path for the UI downloader
    return str(output_path), slide_counter


""" 
def run_generator_v1(): 
    # User Inputs
    raw_pptx = input("Paste/Enter template PPTX path: ").strip().replace('"', '')
    raw_excel = input("Paste/Enter mapping Excel path: ").strip().replace('"', '')
    
    template_path = Path(raw_pptx)
    excel_path = Path(raw_excel)
    output_path = template_path.parent / (template_path.stem + "_v1_generated.pptx")

    # Load Data
    wb = load_workbook(str(excel_path), data_only=True)
    ws = wb["Content"]
    tags = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    max_col = ws.max_column

    # Start PowerPoint
    try:
        ppt_app = win32com.client.GetActiveObject("PowerPoint.Application")
    except:
        ppt_app = win32com.client.Dispatch("PowerPoint.Application")
    
    prs = ppt_app.Presentations.Open(str(template_path.absolute()), WithWindow=True)
    source_slide = prs.Slides(1)
    
    # Track the insertion point (start after the template)
    current_pos = 2 

    # Generation Loop
    for col_idx in range(3, max_col + 1):
        scorecard_name = ws.cell(row=1, column=col_idx).value
        print(f"Generating: {scorecard_name}")

        new_slide = source_slide.Duplicate()
        # Ensure slide is moved to the end to maintain Excel order
        new_slide.MoveTo(prs.Slides.Count)
        
        # Collect column data
        column_data = {}
        for row_idx, tag in enumerate(tags, start=2):
            column_data[tag] = ws.cell(row=row_idx, column=col_idx).value

        # Update Shapes
        for shape in new_slide.Shapes:
            # Tag Matching Logic (Simplified)
            s_name = shape.Name.replace(" ", "_")
            if shape.HasTable: kind = "Table"
            elif shape.HasTextFrame:
                kind = "Label" if len(shape.TextFrame.TextRange.Text.strip()) < 25 else "TextBox"
            else: kind = "Shape"
            tag = f"{kind}_{s_name}"

            if tag in column_data and column_data[tag] is not None:
                data_val = column_data[tag]
                
                try:
                    if shape.HasTextFrame:
                        shape.TextFrame.TextRange.Text = excel_to_pptx_text(data_val)
                    elif shape.HasTable:
                        update_table_v1(shape.Table, data_val)
                except Exception as e:
                    print(f"  Error on {tag}: {e}")

    prs.SaveAs(str(output_path.absolute()))
    print(f"\nSuccess! Version 1.0 saved at: {output_path}")
    """

if __name__ == "__main__":
    p_path = input("Template Path: ").strip().replace('"', '')
    e_path = input("Excel Path: ").strip().replace('"', '')
    result = generate_deck(p_path, e_path)
    print(f"Generated: {result}")